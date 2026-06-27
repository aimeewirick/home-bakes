"""
generate_recipe_template.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Connects to Firestore and generates an Excel recipe
upload template with pre-loaded dropdowns for:
  - Meal Type
  - Recipe Category
  - Unit (per ingredient)
  - Is Public (true/false)

HOW TO USE:
  1. Run: python generate_recipe_template.py
  2. Open the generated recipes_upload_template.xlsx
  3. Fill in your recipes (up to 50 rows pre-formatted)
  4. Save As → CSV (Comma delimited) when ready to upload
  5. Run: python upload_recipes.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import os
import json
import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Firebase init ─────────────────────────────────────────────────────────────
print("\n🔥 Connecting to Firebase...")
cred_json = os.environ.get("FIREBASE_CREDENTIALS")
if cred_json:
    cred = credentials.Certificate(json.loads(cred_json))
else:
    cred = credentials.Certificate("firebase_admin_key.json")

if not firebase_admin._apps:
    firebase_admin.initialize_app(cred)

db = firestore.client()
print("✅ Connected!\n")

# ── Fetch reference data from Firestore ───────────────────────────────────────
print("📂 Fetching reference data...")

meal_types = sorted(
    [d.to_dict()["name"] for d in db.collection("meal_types").stream()
     if d.to_dict().get("name") != "All"]
)

recipe_categories = sorted(
    [d.to_dict()["name"] for d in db.collection("recipe_categories").stream()
     if d.to_dict().get("name") != "All"]
)

units = [d.to_dict() for d in db.collection("units").stream()]
unit_names = sorted([u["name"] for u in units])

print(f"  ✅ {len(meal_types)} meal types")
print(f"  ✅ {len(recipe_categories)} recipe categories")
print(f"  ✅ {len(unit_names)} units\n")

# ── Build Excel workbook ──────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Hidden reference sheet for dropdown lists ─────────────────────────────────
ref_sheet = wb.create_sheet("_ref")
ref_sheet.sheet_state = "hidden"

# Write meal types
for i, name in enumerate(meal_types, start=1):
    ref_sheet.cell(row=i, column=1, value=name)

# Write recipe categories
for i, name in enumerate(recipe_categories, start=1):
    ref_sheet.cell(row=i, column=2, value=name)

# Write units
for i, name in enumerate(unit_names, start=1):
    ref_sheet.cell(row=i, column=3, value=name)

# ── Main recipe sheet ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Recipes"

# ── Styles ────────────────────────────────────────────────────────────────────
TEAL        = "4A9FA5"
TEAL_LIGHT  = "E0F2F3"
CREAM       = "FAF7F0"
DARK        = "2C2C2C"
GREY        = "888880"
RED         = "C0392B"

header_fill    = PatternFill("solid", fgColor=TEAL)
required_fill  = PatternFill("solid", fgColor="FFE8E8")
optional_fill  = PatternFill("solid", fgColor=TEAL_LIGHT)
example_fill   = PatternFill("solid", fgColor=CREAM)
header_font    = Font(bold=True, color="FFFFFF", size=10)
required_font  = Font(bold=True, color=RED, size=9)
optional_font  = Font(bold=True, color=DARK, size=9)
example_font   = Font(italic=True, color=GREY, size=9)
thin_border    = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD")
)

# ── Column definitions ────────────────────────────────────────────────────────
# (header, width, required, notes)
columns = [
    ("title",        30, True,  "Recipe name"),
    ("meal_type",    15, True,  "Select from dropdown"),
    ("category",     20, True,  "Select from dropdown"),
    ("servings",     10, False, "Number (e.g. 8)"),
    ("is_public",    10, False, "true or false"),
    ("notes",        30, False, "Optional notes"),
    ("ingredients",  50, True,  "name:qty:unit|name:qty:unit  (pipe separated)"),
    ("directions",   60, True,  "Step 1 text|Step 2 text  (pipe separated)"),
]

# ── Title banner ──────────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "🏠 HomeBakes — Recipe Upload Template"
title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
title_cell.fill  = PatternFill("solid", fgColor=TEAL)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Instructions row ──────────────────────────────────────────────────────────
ws.merge_cells("A2:H2")
inst_cell = ws["A2"]
inst_cell.value = (
    "Fill in recipes below. Red columns = required. "
    "Ingredients: name:quantity:unit separated by |    "
    "Directions: each step separated by |    "
    "Save as CSV when ready to upload."
)
inst_cell.font      = Font(italic=True, size=9, color=GREY)
inst_cell.alignment = Alignment(horizontal="center", vertical="center")
inst_cell.fill      = PatternFill("solid", fgColor=CREAM)
ws.row_dimensions[2].height = 20

# ── Column headers ────────────────────────────────────────────────────────────
for col_idx, (header, width, required, notes) in enumerate(columns, start=1):
    col_letter = get_column_letter(col_idx)

    # Header row (row 3)
    cell = ws.cell(row=3, column=col_idx, value=header.upper())
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border
    ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 22

    # Notes row (row 4)
    note_cell = ws.cell(row=4, column=col_idx, value=f"{'* Required' if required else 'Optional'} — {notes}")
    note_cell.font      = required_font if required else optional_font
    note_cell.fill      = required_fill if required else optional_fill
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.border    = thin_border
    ws.row_dimensions[4].height = 30

# ── Example row (row 5) ───────────────────────────────────────────────────────
examples = [
    "Grandma's Yeast Bread",
    "Dinner",
    "Breads & Cereals",
    "12",
    "true",
    "A family favorite",
    "flour:4:cup|water:1.5:cup|salt:1:tsp|yeast:2:tsp",
    "Mix flour salt and yeast|Add warm water and stir|Knead for 10 minutes|Let rise 1 hour|Bake at 350 for 30 minutes"
]
for col_idx, value in enumerate(examples, start=1):
    cell = ws.cell(row=5, column=col_idx, value=value)
    cell.font      = example_font
    cell.fill      = example_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = thin_border
ws.row_dimensions[5].height = 40

# ── Data rows (6 to 55) ───────────────────────────────────────────────────────
DATA_START = 6
DATA_END   = 55
NUM_ROWS   = DATA_END - DATA_START + 1

# Dropdown ranges on _ref sheet
meal_range = f"_ref!$A$1:$A${len(meal_types)}"
cat_range  = f"_ref!$B$1:$B${len(recipe_categories)}"
unit_range = f"_ref!$C$1:$C${len(unit_names)}"

# Meal type dropdown — column B
dv_meal = DataValidation(
    type="list",
    formula1=meal_range,
    allow_blank=True,
    showDropDown=False
)
dv_meal.sqref = f"B{DATA_START}:B{DATA_END}"
ws.add_data_validation(dv_meal)

# Category dropdown — column C
dv_cat = DataValidation(
    type="list",
    formula1=cat_range,
    allow_blank=True,
    showDropDown=False
)
dv_cat.sqref = f"C{DATA_START}:C{DATA_END}"
ws.add_data_validation(dv_cat)

# is_public dropdown — column E
dv_public = DataValidation(
    type="list",
    formula1='"true,false"',
    allow_blank=True,
    showDropDown=False
)
dv_public.sqref = f"E{DATA_START}:E{DATA_END}"
ws.add_data_validation(dv_public)

# Style data rows
for row in range(DATA_START, DATA_END + 1):
    ws.row_dimensions[row].height = 18
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border    = thin_border
        # Subtle alternating row color
        if row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F8F5EE")

# Freeze panes below header
ws.freeze_panes = "A6"

# ── Save ──────────────────────────────────────────────────────────────────────
output_file = "recipes_upload_template.xlsx"
wb.save(output_file)
print(f"✅ Template saved as: {output_file}")
print(f"\n📋 Instructions:")
print(f"   1. Open {output_file} in Excel")
print(f"   2. Fill in your recipes starting at row 6")
print(f"   3. Use dropdowns for Meal Type, Category, and Is Public")
print(f"   4. Ingredients: flour:2:cup|sugar:1:tbsp|eggs:2:each")
print(f"   5. Directions:  Mix dry ingredients|Add wet ingredients|Bake 30 min")
print(f"   6. File → Save As → CSV (Comma delimited)")
print(f"   7. Run: python upload_recipes.py\n")
